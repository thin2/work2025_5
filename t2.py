import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import os

EXCEL_FILE = 'books.xlsx'

class Book:
    def __init__(self, name, author, year, add_time=None, status='未借出', borrower='', borrow_date='', return_date=''):
        self.name = name
        self.author = author
        self.year = year
        self.add_time = add_time or datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.status = status
        self.borrower = borrower
        self.borrow_date = borrow_date
        self.return_date = return_date

    def to_list(self):
        return [self.name, self.author, self.year, self.add_time,
                self.status, self.borrower, self.borrow_date, self.return_date]

class Library:
    def __init__(self):
        self.books = []
        self.load_books()

    def load_books(self):
        if not os.path.exists(EXCEL_FILE):
            self.save_books()
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            book = Book(*row)
            self.books.append(book)
        wb.close()

    def save_books(self):
        wb = Workbook()
        ws = wb.active
        ws.append(['书名', '作者', '出版年份', '添加时间', '借阅状态', '借阅人', '借阅时间', '应归还时间'])
        for book in self.books:
            ws.append(book.to_list())
        wb.save(EXCEL_FILE)

    def add_book(self):
        name = input("书名：")
        author = input("作者：")
        year = input("出版年份：")
        book = Book(name, author, year)
        self.books.append(book)
        self.save_books()
        print("✅ 图书添加成功！")

    def display_books(self):
        if not self.books:
            print("暂无图书记录。")
            return
        for book in self.books:
            print(f"书名: {book.name}, 作者: {book.author}, 出版年份: {book.year}, 添加时间: {book.add_time}, "
                  f"状态: {book.status}, 借阅人: {book.borrower}, 借出: {book.borrow_date}, 应还: {book.return_date}")

    def find_book_by_name(self):
        name = input("请输入书名关键词：")
        found = False
        for book in self.books:
            if name.lower() in book.name.lower():
                print(f"书名: {book.name}, 作者: {book.author}, 状态: {book.status}")
                found = True
        if not found:
            print("未找到相关图书。")

    def delete_book(self):
        name = input("请输入要删除的图书名称：")
        for book in self.books:
            if book.name == name:
                if book.status == '已借出':
                    print("❌ 无法删除，图书已借出。")
                    return
                self.books.remove(book)
                self.save_books()
                print("✅ 图书删除成功。")
                return
        print("未找到该图书。")

    def borrow_book(self):
        name = input("请输入要借阅的图书名称：")
        for book in self.books:
            if book.name == name and book.status == '未借出':
                borrower = input("请输入借阅人姓名：")
                borrow_date = datetime.now().strftime('%Y-%m-%d')
                return_date = (datetime.now() + timedelta(days=14)).strftime('%Y-%m-%d')
                book.status = '已借出'
                book.borrower = borrower
                book.borrow_date = borrow_date
                book.return_date = return_date
                self.save_books()
                print(f"✅ 借阅成功，应在 {return_date} 前归还。")
                return
        print("❌ 图书不存在或已被借出。")

    def return_book(self):
        name = input("请输入归还的图书名称：")
        for book in self.books:
            if book.name == name and book.status == '已借出':
                book.status = '未借出'
                book.borrower = ''
                book.borrow_date = ''
                book.return_date = ''
                self.save_books()
                print("✅ 图书归还成功。")
                return
        print("❌ 未找到借出的该图书。")

    def sort_by_name(self):
        sorted_books = sorted(self.books, key=lambda x: x.name)
        for book in sorted_books:
            print(f"{book.name} - {book.author}")

    def sort_by_add_time(self):
        sorted_books = sorted(self.books, key=lambda x: x.add_time)
        for book in sorted_books:
            print(f"{book.add_time} - {book.name}")

    def menu(self):
        while True:
            print("\n📚 图书管理系统菜单")
            print("1. 添加图书")
            print("2. 显示所有图书")
            print("3. 按书名查找图书")
            print("4. 删除图书")
            print("5. 借阅图书")
            print("6. 归还图书")
            print("7. 按书名排序图书")
            print("8. 按添加时间排序图书")
            print("9. 退出系统")
            choice = input("请输入操作编号：")
            try:
                if choice == '1':
                    self.add_book()
                elif choice == '2':
                    self.display_books()
                elif choice == '3':
                    self.find_book_by_name()
                elif choice == '4':
                    self.delete_book()
                elif choice == '5':
                    self.borrow_book()
                elif choice == '6':
                    self.return_book()
                elif choice == '7':
                    self.sort_by_name()
                elif choice == '8':
                    self.sort_by_add_time()
                elif choice == '9':
                    self.save_books()
                    print("📁 数据已保存，系统退出。")
                    break
                else:
                    print("❌ 输入无效，请重新输入。")
            except Exception as e:
                print(f"⚠️ 操作出错：{e}")

if __name__ == '__main__':
    library = Library()
    library.menu()
