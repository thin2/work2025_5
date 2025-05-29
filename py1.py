import pandas as pd
import os

# 读取原始资产数据 Excel 文件
df = pd.read_excel('Excel设备统计表素材.xlsx')

# 创建输出文件夹用于存放按“保管人”拆分的文件
output_dir = "保管人拆分表"
os.makedirs(output_dir, exist_ok=True)

# 按“保管人”列分组，并分别保存为 Excel 文件
for keeper, group in df.groupby("保管人"):
    filename = os.path.join(output_dir, f"{keeper}.xlsx")
    group.to_excel(filename, index=False)

# 以下为将修改后的文件合并，并标注出修改或新增的部分
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

original_file = "Excel设备统计表素材.xlsx"
edited_folder = "保管人拆分表"
output_file = "新资产总表.xlsx"

# 读取原始资产数据，所有列转换为字符串类型，填充空值为""
original_df = pd.read_excel(original_file, dtype=str)
original_df.fillna("", inplace=True)

# 将原始数据构建为字典，key为资产编号
original_dict = {row["资产编号"]: row for _, row in original_df.iterrows()}

new_rows = []       # 存储新增的数据行
updated_rows = {}   # 存储被修改的数据行及其变更列
column = "备注"      # 用于标记备注列

# 遍历拆分后的所有文件
for filename in os.listdir(edited_folder):
    if filename.endswith(".xlsx"):
        edited_path = os.path.join(edited_folder, filename)
        edited_df = pd.read_excel(edited_path, dtype=str)
        edited_df.fillna("", inplace=True)

        # 遍历每一行数据，比较与原始数据的差异
        for _, edited_row in edited_df.iterrows():
            zcbh = str(edited_row["资产编号"])
            if zcbh in original_dict:
                original_row = original_dict[zcbh]
                changed_cols = []

                # 检查每一列是否有修改
                for col in original_df.columns:
                    if col != "资产编号" and col in edited_row and edited_row[col] != original_row[col]:
                        changed_cols.append(col)

                # 如果有修改，更新原始记录，标记为“修改”
                if changed_cols:
                    edited_row[column] = "修改"
                    updated_rows[zcbh] = (edited_row, changed_cols)
                    original_dict[zcbh] = edited_row
            else:
                # 如果资产编号不存在于原始记录中，标记为“新增”
                edited_row[column] = "新增"
                new_rows.append(edited_row)

# 构建最终的数据列表，包括更新后的原始数据和新增数据
final_rows = []
change_tracking = {}  # 用于记录哪些资产编号被修改了哪些列

for _, row in original_df.iterrows():
    zcbh = row["资产编号"]
    if zcbh in original_dict:
        final_rows.append(original_dict[zcbh])
        if zcbh in updated_rows:
            change_tracking[zcbh] = updated_rows[zcbh][1]

# 添加新增行
for new_row in new_rows:
    final_rows.append(new_row)
    change_tracking[new_row["资产编号"]] = "新增"

# 构建新的 DataFrame 并输出到 Excel 文件
final_df = pd.DataFrame(final_rows, columns=original_df.columns)
final_df.to_excel(output_file, index=False)

# 用 openpyxl 给被修改或新增的单元格标黄
wb = load_workbook(output_file)
ws = wb.active
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 获取表头和对应列索引
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col_index_map = {col: idx + 1 for idx, col in enumerate(headers)}

# 遍历每一行，根据变更记录给相应单元格上色
for row in ws.iter_rows(min_row=2):
    zcbh = str(row[col_index_map["资产编号"] - 1].value)
    if zcbh in change_tracking:
        changes = change_tracking[zcbh]

        if changes == "新增":
            # 新增行整体标黄
            for cell in row:
                cell.fill = yellow_fill
        else:
            # 仅修改的列标黄
            for changed_col in changes:
                col_idx = col_index_map.get(changed_col)
                if col_idx:
                    row[col_idx - 1].fill = yellow_fill

# 保存最终文件
wb.save(output_file)
print(f"合并完成，修改单元格已标黄，保存为：{output_file}")
