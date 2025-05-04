import tkinter as tk
from tkinter import messagebox
import pymysql


def get_connection():
    """
    建立并返回一个与 MySQL 数据库的连接。
    注意：请根据实际情况修改 host、user、password、database 等参数。
    """
    return pymysql.connect(
        host="localhost",
        user="root",
        password="h041220H",
        database="student_score_db",
        charset="utf8mb4"
    )


class MainApp:
    """
    主应用程序类，负责显示教师登录窗口，并进行登录验证。
    """

    def __init__(self, master):
        """
        初始化登录界面。
        参数：
            master - Tkinter 根窗口对象。
        """
        self.master = master
        self.master.title("学生成绩管理系统")  # 设置窗口标题

        # 显示登录标题
        tk.Label(master, text="教师登录").grid(row=0, column=0, columnspan=2, pady=10)

        # 显示用户名标签和输入框
        tk.Label(master, text="用户名:").grid(row=1, column=0, padx=5, pady=5)
        self.username_entry = tk.Entry(master)
        self.username_entry.grid(row=1, column=1, padx=5, pady=5)

        # 显示密码标签和输入框，密码以“*”显示
        tk.Label(master, text="密码:").grid(row=2, column=0, padx=5, pady=5)
        self.password_entry = tk.Entry(master, show="*")
        self.password_entry.grid(row=2, column=1, padx=5, pady=5)

        # 登录按钮，点击后调用 login 方法
        tk.Button(master, text="登录", command=self.login).grid(row=3, column=0, columnspan=2, pady=10)

    def login(self):
        """
        登录验证函数。
        获取用户名和密码后，与数据库中的教师登录表 teacherLogin 进行比对，
        登录成功后打开教师管理窗口，并隐藏登录窗口。
        """
        username = self.username_entry.get()  # 获取输入的用户名
        password = self.password_entry.get()  # 获取输入的密码

        # 使用数据库连接进行验证
        with get_connection() as conn:
            cursor = conn.cursor()
            sql = "SELECT * FROM teacherLogin WHERE username=%s AND password=%s"
            cursor.execute(sql, (username, password))
            result = cursor.fetchone()

            if result:
                # 如果查到记录，说明验证成功
                messagebox.showinfo("提示", "登录成功！")
                # 打开教师管理功能窗口（新窗口使用 Toplevel 创建）
                TeacherWindow(tk.Toplevel(self.master))
                # 隐藏当前登录窗口，避免整个程序退出
                self.master.withdraw()
            else:
                # 登录失败，弹出错误提示
                messagebox.showerror("错误", "用户名或密码错误")


class TeacherWindow:
    """
    教师管理窗口，负责显示学生成绩的管理功能。
    包含查询、添加、更新、删除学生记录，以及 Excel 导入/导出功能。
    """

    def __init__(self, master):
        """
        初始化教师管理窗口。
        参数：
            master - Toplevel 窗口对象。
        """
        self.master = master
        self.master.title("教师管理")

        # 绑定窗口关闭事件，点击右上角×时调用 on_closing 方法退出整个系统
        self.master.protocol("WM_DELETE_WINDOW", self.on_closing)

        # 创建功能按钮，每个按钮调用对应的功能方法
        tk.Button(master, text="查询所有学生", command=self.query_students).pack(pady=5)
        tk.Button(master, text="添加学生", command=self.add_student).pack(pady=5)
        tk.Button(master, text="更新学生成绩", command=self.update_student).pack(pady=5)
        tk.Button(master, text="删除学生", command=self.delete_student).pack(pady=5)
        tk.Button(master, text="导入Excel", command=self.import_from_excel).pack(pady=5)
        tk.Button(master, text="导出Excel", command=self.export_to_excel).pack(pady=5)

    def on_closing(self):
        """
        处理教师管理窗口的关闭事件。
        关闭窗口后，退出整个程序。
        """
        self.master.destroy()  # 关闭教师管理窗口
        import sys
        sys.exit()  # 退出程序

    def query_students(self):
        """
        查询数据库中的所有学生记录，并以弹窗的形式显示。
        以制表符分隔学号、姓名和三门课成绩。
        """
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT stu_no, name, chinese, math, english FROM student")
            results = cursor.fetchall()

            # 构建显示信息，第一行为表头
            info = "学号\t姓名\t语文\t数学\t英语\n"
            for row in results:
                info += f"{row[0]}\t{row[1]}\t{row[2]}\t{row[3]}\t{row[4]}\n"

            messagebox.showinfo("所有学生成绩", info)

    def add_student(self):
        """
        弹出新窗口输入学生信息（学号、姓名、语文、数学、英语成绩），
        然后将数据插入到 student 数据库表中。
        """

        def save():
            # 获取输入的学生信息
            stu_no = e_stu_no.get()
            name = e_name.get()
            chinese = e_chinese.get()
            math = e_math.get()
            english = e_english.get()

            # 将数据插入到数据库中
            with get_connection() as conn:
                cursor = conn.cursor()
                sql = "INSERT INTO student (stu_no, name, chinese, math, english) VALUES (%s, %s, %s, %s, %s)"
                cursor.execute(sql, (stu_no, name, chinese, math, english))
                conn.commit()
            # 弹出提示信息，提示添加成功，并关闭添加窗口
            messagebox.showinfo("提示", "添加成功")
            win_add.destroy()

        # 创建添加学生的弹窗
        win_add = tk.Toplevel(self.master)
        win_add.title("添加学生")

        # 标签和输入框：学号
        tk.Label(win_add, text="学号:").grid(row=0, column=0, padx=5, pady=5)
        e_stu_no = tk.Entry(win_add)
        e_stu_no.grid(row=0, column=1, padx=5, pady=5)

        # 标签和输入框：姓名
        tk.Label(win_add, text="姓名:").grid(row=1, column=0, padx=5, pady=5)
        e_name = tk.Entry(win_add)
        e_name.grid(row=1, column=1, padx=5, pady=5)

        # 标签和输入框：语文成绩
        tk.Label(win_add, text="语文:").grid(row=2, column=0, padx=5, pady=5)
        e_chinese = tk.Entry(win_add)
        e_chinese.grid(row=2, column=1, padx=5, pady=5)

        # 标签和输入框：数学成绩
        tk.Label(win_add, text="数学:").grid(row=3, column=0, padx=5, pady=5)
        e_math = tk.Entry(win_add)
        e_math.grid(row=3, column=1, padx=5, pady=5)

        # 标签和输入框：英语成绩
        tk.Label(win_add, text="英语:").grid(row=4, column=0, padx=5, pady=5)
        e_english = tk.Entry(win_add)
        e_english.grid(row=4, column=1, padx=5, pady=5)

        # 保存按钮，点击后调用内部函数 save
        tk.Button(win_add, text="保存", command=save).grid(row=5, column=0, columnspan=2, pady=10)

    def update_student(self):
        """
        弹出新窗口，根据输入的学号更新对应学生的成绩。
        只更新语文、数学和英语成绩。
        """

        def save_update():
            # 获取输入的学号和新成绩
            stu_no = e_stu_no.get()
            chinese = e_chinese.get()
            math = e_math.get()
            english = e_english.get()

            # 执行更新操作
            with get_connection() as conn:
                cursor = conn.cursor()
                sql = "UPDATE student SET chinese=%s, math=%s, english=%s WHERE stu_no=%s"
                cursor.execute(sql, (chinese, math, english, stu_no))
                conn.commit()

            # 弹出提示，更新成功后关闭更新窗口
            messagebox.showinfo("提示", "更新成功")
            win_update.destroy()

        # 创建更新学生成绩的弹窗
        win_update = tk.Toplevel(self.master)
        win_update.title("更新学生成绩")

        # 标签和输入框：学号
        tk.Label(win_update, text="学号:").grid(row=0, column=0, padx=5, pady=5)
        e_stu_no = tk.Entry(win_update)
        e_stu_no.grid(row=0, column=1, padx=5, pady=5)

        # 标签和输入框：新语文成绩
        tk.Label(win_update, text="新语文成绩:").grid(row=1, column=0, padx=5, pady=5)
        e_chinese = tk.Entry(win_update)
        e_chinese.grid(row=1, column=1, padx=5, pady=5)

        # 标签和输入框：新数学成绩
        tk.Label(win_update, text="新数学成绩:").grid(row=2, column=0, padx=5, pady=5)
        e_math = tk.Entry(win_update)
        e_math.grid(row=2, column=1, padx=5, pady=5)

        # 标签和输入框：新英语成绩
        tk.Label(win_update, text="新英语成绩:").grid(row=3, column=0, padx=5, pady=5)
        e_english = tk.Entry(win_update)
        e_english.grid(row=3, column=1, padx=5, pady=5)

        # 保存按钮，点击后调用内部函数 save_update
        tk.Button(win_update, text="保存", command=save_update).grid(row=4, column=0, columnspan=2, pady=10)

    def delete_student(self):
        """
        弹出新窗口，根据输入的学号删除对应学生记录。
        """

        def confirm_delete():
            # 获取输入的学号
            stu_no = e_stu_no.get()
            # 执行删除操作
            with get_connection() as conn:
                cursor = conn.cursor()
                sql = "DELETE FROM student WHERE stu_no=%s"
                cursor.execute(sql, (stu_no,))
                conn.commit()
            # 弹出提示，删除成功后关闭删除窗口
            messagebox.showinfo("提示", "删除成功")
            win_del.destroy()

        # 创建删除学生记录的弹窗
        win_del = tk.Toplevel(self.master)
        win_del.title("删除学生")

        # 标签和输入框：学号
        tk.Label(win_del, text="学号:").grid(row=0, column=0, padx=5, pady=5)
        e_stu_no = tk.Entry(win_del)
        e_stu_no.grid(row=0, column=1, padx=5, pady=5)

        # 删除按钮，点击后调用内部函数 confirm_delete
        tk.Button(win_del, text="删除", command=confirm_delete).grid(row=1, column=0, columnspan=2, pady=10)

    def import_from_excel(self):
        """
        从 Excel 文件中导入学生成绩数据。
        使用 tkinter.filedialog 打开文件选择对话框，读取 Excel 表中第二行开始的数据，
        根据学号是否存在自动更新或插入数据。
        """
        import openpyxl
        from tkinter.filedialog import askopenfilename

        # 弹出文件选择对话框
        file_path = askopenfilename(title="选择Excel文件", filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path:
            return  # 如果未选择文件，则退出该函数

        # 加载 Excel 工作簿，并获取活动工作表
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        row_count = 0  # 用于统计导入记录数
        with get_connection() as conn:
            cursor = conn.cursor()
            # 从第二行开始遍历每一行数据（第一行为表头）
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # 如果学号为空，则跳过该行
                    continue
                # 假设 Excel 列的顺序为：学号、姓名、语文、数学、英语
                stu_no, name, chinese, math, english = row
                # 使用 INSERT ... ON DUPLICATE KEY UPDATE 语句，
                # 如果学号已存在则更新信息，否则插入新记录
                sql = """
                    INSERT INTO student (stu_no, name, chinese, math, english)
                    VALUES (%s, %s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE
                        name=VALUES(name),
                        chinese=VALUES(chinese),
                        math=VALUES(math),
                        english=VALUES(english)
                """
                cursor.execute(sql, (stu_no, name, chinese, math, english))
                row_count += 1
            conn.commit()

        # 显示导入成功提示，并告知导入记录数
        messagebox.showinfo("提示", f"成功导入 {row_count} 条记录")

    def export_to_excel(self):
        """
        将数据库中所有学生成绩导出到 Excel 文件。
        使用 tkinter.filedialog 保存文件对话框选择保存路径，
        并将查询结果写入 Excel 文件中。
        """
        import openpyxl
        from tkinter.filedialog import asksaveasfilename

        # 弹出文件保存对话框，选择保存路径
        file_path = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return

        # 创建新的 Excel 工作簿及工作表
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "学生成绩"

        # 写入表头
        headers = ["stu_no", "name", "chinese", "math", "english"]
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)

        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT stu_no, name, chinese, math, english FROM student")
            results = cursor.fetchall()

            # 将查询到的数据逐行写入 Excel 表中
            row_index = 2  # 从第二行开始写数据
            for row_data in results:
                for col_idx, value in enumerate(row_data, start=1):
                    sheet.cell(row=row_index, column=col_idx, value=value)
                row_index += 1

        # 保存 Excel 文件到指定路径
        wb.save(file_path)
        # 弹出提示，显示导出成功及保存路径
        messagebox.showinfo("提示", f"成功导出到 {file_path}")


def main():
    """
    程序入口函数。
    创建 Tkinter 根窗口，初始化 MainApp 类，并启动主循环。
    """
    root = tk.Tk()
    MainApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
